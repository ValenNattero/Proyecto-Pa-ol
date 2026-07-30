from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from sqlalchemy import or_
import io
import openpyxl

from .. import models, schemas
from ..database import get_db

router = APIRouter(
    prefix="/herramientas",
    tags=["herramientas"],
)

def generar_codigo_incremental(db: Session, categoria=None, codigos_asignados: set = None) -> str:
    if codigos_asignados is None:
        codigos_asignados = set()
        
    herramientas_db = db.query(models.Herramienta).all()
    max_num = 0
    for h in herramientas_db:
        if h.codigo:
            cod_str = str(h.codigo).strip()
            if cod_str.isdigit():
                val = int(cod_str)
                if val > max_num:
                    max_num = val
            else:
                digits = "".join(filter(str.isdigit, cod_str))
                if digits:
                    try:
                        val = int(digits)
                        if val > max_num:
                            max_num = val
                    except ValueError:
                        pass
                
    siguiente_num = max_num + 1
    while True:
        candidato = str(siguiente_num)
        if candidato not in codigos_asignados:
            existe = db.query(models.Herramienta).filter(models.Herramienta.codigo == candidato).first()
            if not existe:
                codigos_asignados.add(candidato)
                return candidato
        siguiente_num += 1

@router.get("/proximo-codigo")
def obtener_proximos_codigos(count: int = Query(10, ge=1, le=100), db: Session = Depends(get_db)):
    codigos_asignados = set()
    resultado = []
    for _ in range(count):
        cod = generar_codigo_incremental(db, None, codigos_asignados)
        resultado.append(cod)
        codigos_asignados.add(cod)
    return {
        "proximo_codigo": resultado[0] if resultado else "1",
        "siguientes": resultado
    }

@router.get("/buscar", response_model=List[schemas.Herramienta])
def buscar_herramientas(
    q: Optional[str] = Query(None, description="Término de búsqueda para código o descripción"),
    db: Session = Depends(get_db)
):
    query = db.query(models.Herramienta)
    if q:
        query = query.filter(
            or_(
                models.Herramienta.codigo.ilike(f"%{q}%"),
                models.Herramienta.descripcion.ilike(f"%{q}%")
            )
        )
    return query.all()

@router.get("/inventario")
def obtener_inventario_completo(db: Session = Depends(get_db)):
    herramientas_db = db.query(models.Herramienta).order_by(models.Herramienta.codigo.asc()).all()
    hubo_cambios = False
    codigos_existentes = set()
    for h in herramientas_db:
        if h.codigo:
            codigos_existentes.add(h.codigo)
            
    for h in herramientas_db:
        if not h.codigo or not str(h.codigo).strip():
            nuevo_codigo = generar_codigo_incremental(db, h.categoria, codigos_existentes)
            h.codigo = nuevo_codigo
            if not h.codigo_qr:
                h.codigo_qr = f"PANOL-QR-{nuevo_codigo}"
            hubo_cambios = True
            
    if hubo_cambios:
        db.commit()
        
    resultado = []
    for h in herramientas_db:
        prestamo_activo = db.query(models.Prestamo).filter(
            models.Prestamo.herramienta_id == h.id,
            models.Prestamo.estado == models.EstadoPrestamo.pendiente
        ).first()

        item = {
            "id": h.id,
            "codigo": str(h.codigo) if h.codigo else str(h.id),
            "descripcion": h.descripcion or "Sin descripción",
            "categoria": h.categoria.value if h.categoria else "Herramienta",
            "marca": h.marca or "-",
            "origen": h.origen or "PMI",
            "estado": h.estado.value if h.estado else "En servicio",
            "fecha_alta": h.fecha_alta.isoformat() if h.fecha_alta else None,
            "disponible_en_panol": prestamo_activo is None,
            "prestado_a": f"{prestamo_activo.nombre_solicitante} {prestamo_activo.apellido_solicitante}" if prestamo_activo else None,
            "cargo_solicitante": prestamo_activo.cargo_solicitante if prestamo_activo else None,
            "fecha_retiro": prestamo_activo.fecha_retiro.isoformat() if (prestamo_activo and prestamo_activo.fecha_retiro) else None
        }
        resultado.append(item)
    return resultado

@router.post("/", response_model=schemas.Herramienta, status_code=status.HTTP_201_CREATED)
def crear_herramienta(herramienta: schemas.HerramientaCreate, db: Session = Depends(get_db)):
    data = herramienta.model_dump()
    if not data.get("codigo") or not str(data.get("codigo")).strip():
        data["codigo"] = generar_codigo_incremental(db, herramienta.categoria)
    else:
        cod_str = str(data["codigo"]).strip()
        existe = db.query(models.Herramienta).filter(models.Herramienta.codigo == cod_str).first()
        if existe:
            raise HTTPException(status_code=400, detail=f"El código '{cod_str}' ya existe en la base de datos. No se pueden cargar herramientas con código repetido.")
        data["codigo"] = cod_str

    if not data.get("codigo_qr"):
        data["codigo_qr"] = f"PANOL-QR-{data['codigo']}"
        
    db_herramienta = models.Herramienta(**data)
    db.add(db_herramienta)
    db.commit()
    db.refresh(db_herramienta)
    return db_herramienta

@router.post("/bulk", response_model=List[schemas.Herramienta], status_code=status.HTTP_201_CREATED)
def crear_herramientas_bulk(herramientas: List[schemas.HerramientaCreate], db: Session = Depends(get_db)):
    creadas = []
    codigos_asignados = set()
    for item in herramientas:
        data = item.model_dump()
        if not data.get("codigo") or not str(data.get("codigo")).strip():
            data["codigo"] = generar_codigo_incremental(db, item.categoria, codigos_asignados)
        else:
            cod_str = str(data["codigo"]).strip()
            if cod_str in codigos_asignados:
                raise HTTPException(status_code=400, detail=f"El código '{cod_str}' está repetido en la lista a cargar. No se admiten códigos repetidos.")
            existe = db.query(models.Herramienta).filter(models.Herramienta.codigo == cod_str).first()
            if existe:
                raise HTTPException(status_code=400, detail=f"El código '{cod_str}' ya existe en la base de datos. No se pueden cargar herramientas con código repetido.")
            codigos_asignados.add(cod_str)
            data["codigo"] = cod_str
            
        if not data.get("codigo_qr"):
            data["codigo_qr"] = f"PANOL-QR-{data['codigo']}"
            
        db_herramienta = models.Herramienta(**data)
        db.add(db_herramienta)
        creadas.append(db_herramienta)
    db.commit()
    for h in creadas:
        db.refresh(h)
    return creadas

@router.get("/plantilla-excel")
def descargar_plantilla_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Carga"
    
    headers = ["codigo", "descripcion", "categoria", "marca", "origen", "estado"]
    ws.append(headers)
    
    ws.append(["HER-001", "Taladro percutor 750W", "herramienta", "Bosch", "PMI", "En servicio"])
    ws.append(["INS-010", "Disco de corte 115mm", "insumo", "Stanley", "Cooperadora", "En servicio"])
    ws.append(["PRO-005", "Antiparras de seguridad", "material de proteccion", "3M", "Escuela", "En servicio"])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_carga_herramientas.xlsx"}
    )

@router.post("/importar-excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe tener formato Excel (.xlsx o .xls)")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="El archivo Excel está vacío o no tiene filas de datos.")
        
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        
        def get_idx(name):
            return headers.index(name) if name in headers else -1
            
        idx_codigo = get_idx("codigo")
        idx_desc = get_idx("descripcion")
        idx_cat = get_idx("categoria")
        idx_marca = get_idx("marca")
        idx_origen = get_idx("origen")
        idx_estado = get_idx("estado")
        
        if idx_desc == -1:
            raise HTTPException(status_code=400, detail="El archivo debe tener al menos una columna llamada 'descripcion'.")
            
        creadas = []
        codigos_asignados = set()
        for row in rows[1:]:
            if not row or not any(row):
                continue
                
            descripcion_val = str(row[idx_desc]).strip() if idx_desc != -1 and row[idx_desc] else ""
            if not descripcion_val:
                continue
                
            codigo_val = str(row[idx_codigo]).strip() if idx_codigo != -1 and row[idx_codigo] else None
            marca_val = str(row[idx_marca]).strip() if idx_marca != -1 and row[idx_marca] else None
            origen_val = str(row[idx_origen]).strip() if idx_origen != -1 and row[idx_origen] else "PMI"
            
            cat_val = str(row[idx_cat]).strip().lower() if idx_cat != -1 and row[idx_cat] else "herramienta"
            try:
                categoria_enum = models.CategoriaHerramienta(cat_val)
            except ValueError:
                categoria_enum = models.CategoriaHerramienta.herramienta
                
            est_val = str(row[idx_estado]).strip() if idx_estado != -1 and row[idx_estado] else "En servicio"
            try:
                estado_enum = models.EstadoHerramienta(est_val)
            except ValueError:
                estado_enum = models.EstadoHerramienta.en_servicio
                
            if not codigo_val:
                codigo_val = generar_codigo_incremental(db, categoria_enum, codigos_asignados)
            else:
                if codigo_val in codigos_asignados:
                    raise HTTPException(status_code=400, detail=f"El código '{codigo_val}' está repetido en el archivo Excel. No se admiten códigos repetidos.")
                existe_db = db.query(models.Herramienta).filter(models.Herramienta.codigo == codigo_val).first()
                if existe_db:
                    raise HTTPException(status_code=400, detail=f"El código '{codigo_val}' ya existe en la base de datos. No se pueden cargar herramientas con código repetido.")
                codigos_asignados.add(codigo_val)
                
            nueva_h = models.Herramienta(
                codigo=codigo_val,
                descripcion=descripcion_val,
                categoria=categoria_enum,
                marca=marca_val,
                origen=origen_val,
                estado=estado_enum,
                codigo_qr=f"PANOL-QR-{codigo_val}"
            )
            db.add(nueva_h)
            creadas.append(nueva_h)
            
        db.commit()
        for h in creadas:
            db.refresh(h)
            
        items_res = [
            {
                "id": h.id,
                "codigo": h.codigo,
                "descripcion": h.descripcion,
                "categoria": h.categoria.value if h.categoria else "Herramienta",
                "marca": h.marca or "-",
                "origen": h.origen or "PMI",
                "estado": h.estado.value if h.estado else "En servicio"
            }
            for h in creadas
        ]
        return {
            "success": True,
            "importadas": len(creadas),
            "mensaje": f"Se importaron {len(creadas)} herramientas correctamente al inventario.",
            "items": items_res
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error procesando el archivo Excel: {str(e)}")

@router.put("/{herramienta_id}", response_model=schemas.Herramienta)
def modificar_herramienta(herramienta_id: int, herramienta: schemas.HerramientaUpdate, db: Session = Depends(get_db)):
    # TODO: Proteger esta ruta con JWT para que solo administradores puedan acceder
    db_herramienta = db.query(models.Herramienta).filter(models.Herramienta.id == herramienta_id).first()
    if not db_herramienta:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")
    
    update_data = herramienta.model_dump(exclude_unset=True)
    if "codigo" in update_data and update_data["codigo"]:
        nuevo_cod = str(update_data["codigo"]).strip()
        if nuevo_cod != db_herramienta.codigo:
            ot = db.query(models.Herramienta).filter(
                models.Herramienta.codigo == nuevo_cod,
                models.Herramienta.id != herramienta_id
            ).first()
            if ot:
                raise HTTPException(status_code=400, detail=f"El código '{nuevo_cod}' ya se encuentra asignado a otra herramienta.")
            update_data["codigo"] = nuevo_cod
            if not db_herramienta.codigo_qr or "PANOL-QR-" in db_herramienta.codigo_qr:
                update_data["codigo_qr"] = f"PANOL-QR-{nuevo_cod}"
                
    for key, value in update_data.items():
        setattr(db_herramienta, key, value)
        
    db.commit()
    db.refresh(db_herramienta)
    return db_herramienta

@router.delete("/{herramienta_id}")
def eliminar_herramienta(herramienta_id: int, db: Session = Depends(get_db)):
    db_herramienta = db.query(models.Herramienta).filter(models.Herramienta.id == herramienta_id).first()
    if not db_herramienta:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")
    
    db.delete(db_herramienta)
    db.commit()
    return {"success": True, "mensaje": "Herramienta eliminada correctamente del inventario"}
