from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font
import io

from .. import models
from ..database import get_db
from ..security import get_current_token_payload

router = APIRouter(
    prefix="/informes",
    tags=["informes"],
)

@router.get("/generar")
def generar_informe(
    db: Session = Depends(get_db),
    # token_payload: dict = Depends(get_current_token_payload) # Si se requiere auth para descargar
):
    wb = openpyxl.Workbook()
    
    # -------------------------
    # Hoja 1: Préstamos
    # -------------------------
    ws_prestamos = wb.active
    ws_prestamos.title = "Préstamos"
    
    # Encabezados
    headers_prestamos = ["Nombre y Apellido", "Herramientas Solicitadas"]
    ws_prestamos.append(headers_prestamos)
    for cell in ws_prestamos[1]:
        cell.font = Font(bold=True)
        
    # Agrupar préstamos por persona
    # Para simplificar, agruparemos por nombre+apellido
    prestamos_db = db.query(models.Prestamo).all()
    agrupados = {}
    for p in prestamos_db:
        key = f"{p.nombre_panolero} {p.apellido_panolero}"
        desc = p.herramienta.descripcion if p.herramienta else "Desconocida"
        if key not in agrupados:
            agrupados[key] = []
        agrupados[key].append(desc)
        
    for persona, herramientas in agrupados.items():
        ws_prestamos.append([persona, ", ".join(herramientas)])
        
    # -------------------------
    # Hoja 2: Inventario
    # -------------------------
    ws_inventario = wb.create_sheet(title="Inventario Completo")
    
    headers_inv = [
        "Fecha de alta", "Fecha de baja", "ID", "Herramienta/Descripción", 
        "Estado", "Categoría", "Marca", "Origen", "Código QR"
    ]
    ws_inventario.append(headers_inv)
    for cell in ws_inventario[1]:
        cell.font = Font(bold=True)
        
    herramientas_db = db.query(models.Herramienta).all()
    for h in herramientas_db:
        ws_inventario.append([
            h.fecha_alta.strftime("%Y-%m-%d %H:%M:%S") if h.fecha_alta else "",
            h.fecha_baja.strftime("%Y-%m-%d %H:%M:%S") if h.fecha_baja else "",
            h.id,
            h.descripcion,
            h.estado.value if h.estado else "",
            h.categoria.value if h.categoria else "",
            h.marca or "",
            h.origen or "",
            h.codigo_qr or ""
        ])
        
    # Guardar en memoria
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Devolver como archivo descargable
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario.xlsx"}
    )

@router.get("/inventario/excel")
def descargar_excel_inventario(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Pañol"

    # Estilos del encabezado
    headers = [
        "Código", "Descripción", "Categoría", "Marca", 
        "Estado Técnico", "Disponibilidad Actual", "Origen / PMI", "Fecha Alta"
    ]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid")

    herramientas_db = db.query(models.Herramienta).order_by(models.Herramienta.codigo.asc()).all()
    for h in herramientas_db:
        # Verificar si está prestada en este momento
        prestamo_activo = db.query(models.Prestamo).filter(
            models.Prestamo.herramienta_id == h.id,
            models.Prestamo.estado == models.EstadoPrestamo.pendiente
        ).first()

        if prestamo_activo:
            disponibilidad = f"PRESTADA A: {prestamo_activo.nombre_solicitante} {prestamo_activo.apellido_solicitante} ({prestamo_activo.cargo_solicitante})"
        else:
            disponibilidad = "DISPONIBLE EN PAÑOL"

        ws.append([
            h.codigo or str(h.id),
            h.descripcion or "",
            h.categoria.value if h.categoria else "",
            h.marca or "-",
            h.estado.value if h.estado else "En servicio",
            disponibilidad,
            h.origen or "PMI",
            h.fecha_alta.strftime("%d/%m/%Y") if h.fecha_alta else ""
        ])

    # Ajustar ancho de las columnas automáticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Inventario_Panol_Escolar.xlsx"}
    )
